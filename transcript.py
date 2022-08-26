from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl import Workbook
import wget
import os
from functools import cmp_to_key


# 計算該數字在哪個組內
def interval(val):
    if val == 100:
        return 0
    elif val >= 90:
        return 1
    elif val >= 80:
        return 2
    elif val >= 70:
        return 3
    elif val >= 60:
        return 4
    else:
        return 5

#下載成績單模板
def get_transcript_template():
    if not os.path.exists('template.xlsx'):
        wget.download('https://github.com/ryanovovo/Transcript/raw/master/template.xlsx')
    
    return load_workbook('template.xlsx')

# 自定義sort function
def cmp_score(lhs, rhs):
    return rhs.weightedScoreSum - lhs.weightedScoreSum

# 取得班級前十名的成績
def get_top_ten_score(studentTranscripts):
    ret = []
    for i in range(10):
        ret.append(round(studentTranscripts[i].weightedScoreAvg, 2))
    return ret

# 取得組距表
def get_intervals(studentTranscripts):
    ret = [[0 for x in range(7)] for y in range(6)] 

    for student in studentTranscripts:
        for i in range(0, 7):
            ret[interval(student.score[i])][i] += 1
    return ret

# 取得班級各科平均成績
def get_class_avg_score(studentTranscripts):
    ret = [0, 0, 0, 0, 0, 0, 0]
    for student in studentTranscripts:
        for i in range(0, 7):
            ret[i] += student.score[i]
    for i in range(0, 7):
        ret[i] /= len(studentTranscripts)
        ret[i] = round(ret[i], 2)
    return ret


class Student():
    def __init__(self, studentId, rawClassScore):
        self.seatNumber = studentId 
        self.dataRow = 0
        self.rawData = []
        self.score = []
        self.name = ''
        self.transcript = None
        self.weightedScoreSum = 0.0
        self.weightedScoreAvg = 0.0
        self.rank = 0
        self.diff = 0

        # 根據座號查表取得數據所在的row
        for i in range(1, 50):
            if rawClassScore.cell(i, 1).value == self.seatNumber:
                self.dataRow = i

        # 取得該row內的所有數據
        for cell in rawClassScore[str(self.dataRow)]:
            self.rawData.append(cell.value)

        self.name = self.rawData[1]
        self.diff = self.rawData[12]
        self.get_score()
        
    #取得該學生的成績
    def get_score(self):
        weights = [5, 3 ,4, 3, 1, 1, 1]
        for i in range(2, 9):
            self.score.append(self.rawData[i])
        
        self.weightedScoreSum = 0.0
        for i in range(7):
            self.weightedScoreSum += weights[i] * self.score[i]
        
        self.weightedScoreAvg = self.weightedScoreSum / 18
        self.weightedScoreAvg = round(self.weightedScoreAvg, 2)

        self.score.append(self.weightedScoreSum)
        self.score.append(self.weightedScoreAvg)

    # 生成成績單
    def generate_transcript(self, topTenScore, intervals, classAvgScore):
        self.transcript = get_transcript_template()
        self.transcript.active['B3'] = self.name
        self.transcript.active['A4'] = self.seatNumber
        self.transcript.active['L4'] = self.rank
        self.transcript.active['M4'] = self.diff
    
        # 填入學生成績
        for i in range(9):
            self.transcript.active.cell(4, i+3).value = self.score[i]
        
        # 填入組距表
        for i in range(6):
            for j in range(7):
                self.transcript.active.cell(i+6, j+3).value = intervals[i][j]

        # 填入前十名的加權平均成績
        for i in range(10):
            self.transcript.active.cell(10+i, 13).value = topTenScore[i]
        
        # 填入班級各科平均成績
        for i in range(7):
            self.transcript.active.cell(5, 3+i).value = classAvgScore[i]
    # 將excel表存入資料夾
    def save_xlsx(self, path=''):
        self.transcript.save(path+'/'+str(self.seatNumber)+str(self.name)+'.xlsx')
            
# 讀取原始成績檔
rawClassScore = load_workbook('score.xlsx')['Sheet2']

#初始化資料
seatNumbers = []
studentTranscripts = []
topTenScore = []

# 取得所有學生的座號
for cell in rawClassScore['a']:
    if type(cell.value) == int:
        seatNumbers.append(cell.value)
    
    elif type(cell.value) == None:
        break

# 使用座號生成個人成績資訊
for seatNumber in seatNumbers:
    studentTranscripts.append(Student(seatNumber, rawClassScore))

# 使用加權總分排序成績
studentTranscripts = sorted(studentTranscripts, key=cmp_to_key(cmp_score))

# 填入排名
for i in range(0, len(studentTranscripts)):
    studentTranscripts[i].rank = i+1

# 取得前十名成績, 組聚, 班級各科平均成績
topTenScore = get_top_ten_score(studentTranscripts)
intervals = get_intervals(studentTranscripts)
classAvgScore = get_class_avg_score(studentTranscripts)

# 建立存放成績單的資料夾
if not os.path.exists('transcripts'):
    os.mkdir('transcripts')

# 生成成績單並存入資料夾
for st in studentTranscripts:
    st.generate_transcript(topTenScore, intervals, classAvgScore)
    st.save_xlsx('transcripts')
